from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math

wb=load_workbook('function.xlsx')
ws=wb.active
relleno_naranja=PatternFill(start_color='FFA500',end_color='FFA500',fill_type='solid')
relleno_amarillo=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

def tabla1_E030_10_2(zona):
  cel_1,cel_1.fill,cel_1,cel_1.value=ws[zona],relleno_amarillo,ws[zona].offset(row=0,column=-1),'zona'
  resultado='=IF('+zona+'="Z1",0.45,IF('+zona+'="Z2",0.35,IF('+zona+'="Z3",0.25,0.1)))'
  res_1,res_1.fill,res_1.value,res_1,res_1.value=ws[zona].offset(row=1,column=0),relleno_naranja,resultado,ws[zona].offset(row=1,column=-1),'Zona(g)'

def momentoPositivo(wu,ln,tipoExtremo='monolitico'):
  cel_1,cel_1.fill,cel_1,cel_1.value=ws[wu],relleno_amarillo,ws[wu].offset(row=0,column=-1),'wu'
  cel_2,cel_2.fill,cel_2,cel_2.value=ws[ln],relleno_amarillo,ws[ln].offset(row=0,column=-1),'ln'
  #momentos externo
  if not tipoExtremo=='monolitico':
    #el extremo discontinuo no esta restringido
    tramoExtremo='=('+wu+'*POWER('+ln+',2))/11'
  else:
    #el extremo discontinuo es monolitico con el apoyo
    tramoExtremo='=('+wu+'*POWER('+ln+',2))/14'
  #momentos internos
  tramoInterno = '=('+wu+'*POWER('+ln+',2))/16'
  res_1,res_1.fill,res_1.value,res_1,res_1.value=ws[ln].offset(row=1,column=0),relleno_naranja,tramoExtremo,ws[ln].offset(row=1,column=-1),'Momento externo'
  res_1,res_1.fill,res_1.value,res_1,res_1.value=ws[ln].offset(row=2,column=0),relleno_naranja,tramoInterno,ws[ln].offset(row=2,column=-1),'Momento interno'
  return

def cuadratica(a,b,c):
  """
  Function to solve: ax2+bx+c=0
  """
  x=[0,0]
  x[0]=(-b+(b**2-4*a*c)**0.5)/(2*a)
  x[1]=(-b-(b**2-4*a*c)**0.5)/(2*a)
  return x

def cableUniformLoad(w0,x,y):
  """
  Cable subjected to a uniform distribute load
  
  Args:
  w0:distribuida
  x:longitud
  y:altura

  Returns:
  FH and T
  """
  FH=w0*x**2/(2*y)
  teta=math.atan(w0/FH*x)
  T=FH/math.cos(teta)
  return teta*180/math.pi,FH,T

#aplicando fomulas
tabla1_E030_10_2('B1')
momentoPositivo('B8','B9')

wb.save("function.xlsx")
# frozenset({'DGET', 'MIDB', 'MIN', 'DAYS360', 'EXP', 'VARP', 'ODD', 'COUNTIF', 'GESTEP', 'UPPER', 'BIN2OCT', 'RAND', 'PRICE', 'NORMSINV', 'JIS', 'NOMINAL', 'QUOTIENT', 'FACTDOUBLE', 'CHOOSE', 'IMLOG10', 'DATEVALUE', 'WEEKNUM', 'BINOMDIST', 'NEGBINOMDIST', 'YEARFRAC', 'SUMIFS', 'TIMEVALUE', 'TYPE', 'INDEX', 'DPRODUCT', 'NPER', 'MINA', 'TRANSPOSE', 'CUBEMEMBER', 'SUBSTITUTE', 'MAX', 'ROUNDUP', 'HYPERLINK', 'DISC', 'HEX2BIN', 'ODDFYIELD', 'INDIRECT', 'TODAY', 'PRICEMAT', 'SEARCHB', 'IMSUB', 'HEX2OCT', 'IMCOS', 'CELL', 'IMARGUMENT', 'COVAR', 'MID', 'DEVSQ', 'MIRR', 'YEAR', 'LINEST', 'FTEST', 'ABS', 'IMLOG2', 'MULTINOMIAL', 'CORREL', 'ACOS', 'ROWS', 'LOWER', 'LARGE', 'CUBESETCOUNT', 'SUMX2MY2', 'RANK', 'ROW', 'SERIESSUM', 'ZTEST', 'SUMXMY2', 'CHITEST', 'BESSELK', 'POWER', 'COUNT', 'NORMDIST', 'IMSIN', 'DMAX', 'DATE', 'CONVERT', 'LOGINV', 'STDEVP', 'ISBLANK', 'ASINH', 'FIND', 'ODDFPRICE', 'YIELDMAT', 'IMPOWER', 'BIN2HEX', 'EDATE', 'SUM', 'PERMUT', 'GAMMADIST', 'SUBTOTAL', 'ASC', 'HARMEAN', 'SUMSQ', 'FVSCHEDULE', 'PEARSON', 'FINDB', 'ISREF', 'ROMAN', 'RANDBETWEEN', 'EXACT', 'SLOPE', 'MAXA', 'MDURATION', 'FLOOR', 'AVERAGE', 'SQRTPI', 'FALSE', 'SEARCH', 'ISNUMBER', 'COUPNUM', 'ECMA.CEILING', 'RATE', 'RTD', 'CHIDIST', 'STDEVPA STEYX', 'DB', 'INTRATE', 'HYPGEOMDIST', 'FDIST', 'PROB', 'NETWORKDAYS.INTL', 'ISODD', 'N', 'DATEDIF', 'RSQ', 'MINUTE', 'MROUND', 'IF', 'GETPIVOTDATA', 'CONCATENATE', 'ISERR', 'SIN', 'BETADIST', 'CUMPRINC', 'ISERROR', 'SMALL', 'COUPDAYS', 'DOLLARDE', 'MDETERM', 'DOLLAR', 'ISPMT', 'REPT', 'PERCENTRANK', 'BAHTTEXT', 'WEIBULL', 'IRR', 'AMORLINC', 'COLUMN', 'PRICEDISC', 'FINV', 'TRUNC', 'EXPONDIST', 'DAVERAGE', 'LOGNORMDIST', 'AND', 'DCOUNT', 'ERROR.TYPE', 'FISHER', 'FREQUENCY', 'COSH', 'AVERAGEA', 'PPMT', 'POISSON', 'ISNA', 'SUMIF', 'TBILLYIELD', 'ROUND', 'ERFC', 'DURATION', 'DEC2HEX', 'CRITBINOM', 'ERF', 'SIGN', 'CONFIDENCE', 'DDB', 'ISTEXT', 'CUBEKPIMEMBER', 'TBILLEQ', 'ISLOGICAL', 'COLUMNS', 'FIXED', 'CUBEVALUE', 'FV', 'YIELDDISC', 'COUPDAYBS', 'ODDLYIELD', 'MOD', 'DVARP', 'DCOUNTA', 'FISHERINV', 'PI', 'CHIINV', 'LOG10', 'PHONETIC', 'NOW', 'ISNONTEXT', 'EFFECT', 'BESSELJ', 'REPLACE', 'DEC2BIN', 'TRIMMEAN', 'RECEIVED', 'SYD', 'VAR', 'IMAGINARY', 'MATCH', 'TBILLPRICE', 'IMREAL', 'PV', 'CHAR', 'ATAN2', 'CODE', 'DEC2OCT', 'FACT', 'HOUR', 'TAN', 'GEOMEAN', 'TTEST', 'ACCRINT', 'AVERAGEIF', 'COUPPCD', 'COUNTIFS', 'KURT', 'FORECAST', 'LEN', 'HEX2DEC', 'TANH', 'MODE', 'BESSELY', 'LOGEST', 'VALUE', 'STANDARDIZE', 'EVEN', 'DAY', 'GAMMALN', 'AVEDEV', 'SINH', 'COMPLEX', 'IMSQRT', 'AMORDEGRC', 'LCM', 'AVERAGEIFS', 'TIME', 'GROWTH', 'RIGHT', 'BETAINV', 'OCT2DEC', 'BIN2DEC', 'OFFSET', 'OR', 'TRUE ADDRESS', 'RADIANS', 'QUARTILE', 'IMLN', 'PRODUCT', 'YIELD', 'MEDIAN', 'ROUNDDOWN', 'ACOSH', 'VDB', 'DSTDEV', 'COUNTBLANK', 'SQRT', 'CUBESET', 'RIGHTB', 'IMABS', 'SECOND', 'XIRR', 'PMT', 'VLOOKUP', 'VARPA', 'TINV', 'NORMSDIST', 'ASIN', 'GAMMAINV', 'DOLLARFR', 'LENB', 'DELTA', 'IFERROR', 'IPMT', 'IMSUM', 'AREAS', 'SUMPRODUCT', 'IMPRODUCT', 'COUNTA', 'DSUM', 'IMEXP', 'T', 'OCT2BIN', 'TREND', 'NA', 'VARA', 'PERCENTILE', 'OCT2HEX', 'COUPNCD', 'NOT', 'IMDIV', 'MINVERSE', 'DVAR', 'NORMINV', 'WORKDAY.INTL', 'CUBERANKEDMEMBER', 'LOG', 'NETWORKDAYS', 'DEGREES', 'LN', 'ISO.CEILING', 'HLOOKUP', 'SLN', 'NPV', 'INFO', 'REPLACEB', 'LOOKUP', 'XNPV', 'GCD', 'CLEAN', 'ISEVEN', 'ODDLPRICE', 'COS', 'TEXT', 'COMBIN', 'LEFT', 'STDEV STDEVA', 'CUBEMEMBERPROPERTY', 'SUMX2PY2', 'TRIM', 'TDIST', 'PROPER', 'ATAN', 'CUMIPMT', 'ACCRINTM', 'CEILING', 'LEFTB', 'BESSELI', 'EOMONTH', 'MONTH', 'COUPDAYSNC', 'INTERCEPT', 'DSTDEVP', 'SKEW', 'INT', 'DMIN', 'WEEKDAY', 'IMCONJUGATE', 'WORKDAY ', 'ATANH'})

